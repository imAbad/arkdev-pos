"""Punto 11: exportación a Excel de los reportes ya existentes — mismo
criterio que el resto de la app (una sola función que ya recibe filas ya
calculadas, no reimplementa ninguna agregación). openpyxl es la única
dependencia nueva de todo este punto (confirmado en requirements.txt que
no había ninguna librería de Excel instalada todavía) — sin pandas: no se
necesita nada de lo que pandas trae encima para volcar filas ya
calculadas a una hoja.
"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook


def _cell_value(value):
    # openpyxl no acepta Decimal directo en una celda; date/datetime sí
    # los acepta nativo (los vuelve celdas con formato de fecha real, no
    # texto) — salvo datetime CON tzinfo (ej. CashShift.opened_at/
    # closed_at, USE_TZ=True): Excel no tiene concepto de zona horaria en
    # una celda, openpyxl lo rechaza de plano en vez de truncarlo en
    # silencio. Se convierte a hora local (no UTC crudo, sería confuso
    # para Carlos) y se le quita el tzinfo, mismo criterio que
    # formatDateTime() ya usa en el frontend.
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def _write_sheet(sheet, columns, rows):
    sheet.append([label for label, _ in columns])
    for row in rows:
        sheet.append([_cell_value(row.get(key)) for _, key in columns])

    for column_cells in sheet.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)


def _workbook_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_excel_response(*, filename, columns, rows):
    """`columns`: lista de (encabezado, clave-en-la-fila). `rows`: la
    misma lista de dicts que cada reports.services.* ya devuelve para
    JSON — se reutiliza tal cual, esto solo la vuelca a una hoja."""
    workbook = Workbook()
    _write_sheet(workbook.active, columns, rows)
    return _workbook_response(workbook, filename)


def build_multi_sheet_excel_response(*, filename, sheets):
    """Observación de sesión (reporte de cierre de turno detallado): mismo
    mecanismo que build_excel_response, pero para un reporte con más de
    una tabla (resumen del turno + pagos por método + abonos a crédito) —
    una hoja por sección en vez de forzarlas todas a una sola tabla plana.
    `sheets`: lista de (nombre_hoja, columns, rows)."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, columns, rows in sheets:
        sheet = workbook.create_sheet(title=name[:31])  # límite de Excel para nombres de hoja
        _write_sheet(sheet, columns, rows)
    return _workbook_response(workbook, filename)
