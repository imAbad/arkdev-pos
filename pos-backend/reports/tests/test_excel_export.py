"""Punto 11: exportación a Excel de los 4 reportes existentes
(sales-by-product, inventory-valuation, expired-stock,
cash-shift-closures) — mismo acceso que verlos (IsAdministratorOrSupervisor,
ver ReportsApiPermissionTests), openpyxl es la única dependencia nueva
(no había ninguna librería de Excel instalada antes de este punto)."""
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from audit.models import AuditLog
from catalog.models import InventoryAdjustment
from catalog.services import adjust_batch_stock
from catalog.tests.factories import create_batch, create_product
from reports.excel import build_excel_response, build_multi_sheet_excel_response
from sales.services import close_shift
from sales.tests.factories import create_checkout_context, make_sale
from tenants.models import UserProfile
from tenants.tests.factories import create_user_with_profile


class BuildExcelResponseTests(TestCase):
    def test_produces_a_readable_workbook_with_header_and_rows(self):
        response = build_excel_response(
            filename='prueba.xlsx',
            columns=[('Producto', 'product_name'), ('Cantidad', 'quantity')],
            rows=[{'product_name': 'Yogurt', 'quantity': 5}, {'product_name': 'Refresco', 'quantity': 2}],
        )
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('prueba.xlsx', response['Content-Disposition'])

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(rows[0], ('Producto', 'Cantidad'))
        self.assertEqual(rows[1], ('Yogurt', 5))
        self.assertEqual(rows[2], ('Refresco', 2))

    def test_decimal_values_are_written_as_numbers_not_strings(self):
        response = build_excel_response(
            filename='prueba.xlsx',
            columns=[('Ingreso', 'revenue')],
            rows=[{'revenue': Decimal('120.50')}],
        )
        workbook = load_workbook(BytesIO(response.content))
        value = list(workbook.active.iter_rows(values_only=True))[1][0]
        self.assertEqual(value, 120.50)
        self.assertIsInstance(value, float)


class BuildMultiSheetExcelResponseTests(TestCase):
    def test_writes_one_sheet_per_section_with_its_own_header_and_rows(self):
        response = build_multi_sheet_excel_response(
            filename='detalle.xlsx',
            sheets=[
                ('Resumen', [('Cajero', 'user_email')], [{'user_email': 'cajero@donchuy.test'}]),
                ('Pagos por método', [('Método', 'method_label'), ('Total', 'total')], [
                    {'method_label': 'Efectivo', 'total': Decimal('100.00')},
                ]),
            ],
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['Resumen', 'Pagos por método'])
        self.assertEqual(
            list(workbook['Resumen'].iter_rows(values_only=True)),
            [('Cajero',), ('cajero@donchuy.test',)],
        )
        self.assertEqual(
            list(workbook['Pagos por método'].iter_rows(values_only=True)),
            [('Método', 'Total'), ('Efectivo', 100.00)],
        )


class ReportExcelExportApiTests(APITestCase):
    def setUp(self):
        self.ctx = create_checkout_context()
        self.today = timezone.localdate()
        self.admin, _ = create_user_with_profile(
            'admin@donchuy.test', self.ctx['branch'], role=UserProfile.Role.ADMINISTRADOR,
        )

    def _auth(self, user):
        self.client.force_authenticate(user=user)

    def test_sales_by_product_export_returns_an_xlsx_file(self):
        make_sale(self.ctx['shift'], self.ctx['product'], unit_price=Decimal('10.00'))
        self._auth(self.admin)

        response = self.client.get(
            '/api/v1/reports/sales-by-product/',
            {'date_from': self.today, 'date_to': self.today, 'export': 'xlsx'},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0], ('Producto', 'Categoría', 'Cantidad vendida', 'Ingreso', 'IVA'))
        self.assertEqual(rows[1][0], self.ctx['product'].name)

    def test_sales_by_product_export_uses_cashier_columns_when_grouped_by_cashier(self):
        make_sale(self.ctx['shift'], self.ctx['product'], unit_price=Decimal('10.00'))
        self._auth(self.admin)

        response = self.client.get(
            '/api/v1/reports/sales-by-product/',
            {'date_from': self.today, 'date_to': self.today, 'export': 'xlsx', 'group_by': 'cashier'},
        )
        workbook = load_workbook(BytesIO(response.content))
        header = list(workbook.active.iter_rows(values_only=True))[0]
        self.assertEqual(header, ('Cajero', 'Cantidad vendida', 'Ingreso', 'IVA'))

    def test_inventory_valuation_export_returns_an_xlsx_file(self):
        create_batch(self.ctx['product'], self.ctx['branch'], initial_quantity=10)
        self._auth(self.admin)

        response = self.client.get('/api/v1/reports/inventory-valuation/', {'export': 'xlsx'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))
        header = list(workbook.active.iter_rows(values_only=True))[0]
        self.assertEqual(header, ('Producto', 'Categoría', 'Cantidad', 'Valor'))

    def test_expired_stock_export_returns_an_xlsx_file(self):
        create_batch(
            self.ctx['product'], self.ctx['branch'], initial_quantity=5,
            expiration_date=self.today - timedelta(days=1),
        )
        self._auth(self.admin)

        response = self.client.get('/api/v1/reports/expired-stock/', {'export': 'xlsx'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0], ('Producto', 'Lote', 'Sucursal', 'Caducó', 'Cantidad', 'Valor'))
        self.assertEqual(rows[1][0], self.ctx['product'].name)

    def test_cash_shift_closures_export_returns_an_xlsx_file(self):
        close_shift(shift=self.ctx['shift'], closing_user=self.ctx['user'], actual_closing_balance=Decimal('0'))
        self._auth(self.admin)

        response = self.client.get(
            '/api/v1/reports/cash-shift-closures/',
            {'date_from': self.today, 'date_to': self.today, 'export': 'xlsx'},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))
        header = list(workbook.active.iter_rows(values_only=True))[0]
        self.assertIn('Sucursal', header)
        self.assertIn('Diferencia efectivo', header)

    def test_cash_shift_detail_export_returns_a_multi_sheet_workbook(self):
        make_sale(self.ctx['shift'], self.ctx['product'], unit_price=Decimal('100.00'))
        close_shift(shift=self.ctx['shift'], closing_user=self.ctx['user'], actual_closing_balance=Decimal('0'))
        self._auth(self.admin)

        response = self.client.get(
            '/api/v1/reports/cash-shift-detail/', {'shift': self.ctx['shift'].id, 'export': 'xlsx'},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['Resumen', 'Pagos por método', 'Abonos a crédito'])
        summary_header = list(workbook['Resumen'].iter_rows(values_only=True))[0]
        self.assertIn('Ventas (total)', summary_header)
        payments_rows = list(workbook['Pagos por método'].iter_rows(values_only=True))
        self.assertEqual(payments_rows[1][0], 'Efectivo')

    def test_inventory_adjustments_export_returns_an_xlsx_file(self):
        product = create_product(self.ctx['company'], sku='ADJ-XLSX', requires_batch=True)
        batch = create_batch(product, self.ctx['branch'], initial_quantity=10)
        adjust_batch_stock(
            batch=batch, quantity_delta=-2, reason=InventoryAdjustment.Reason.DAMAGE, actor=self.admin,
        )
        self._auth(self.admin)

        response = self.client.get(
            '/api/v1/reports/inventory-adjustments/', {'date_from': self.today, 'date_to': self.today, 'export': 'xlsx'},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0], ('Producto', 'Lote', 'Sucursal', 'Ajuste', 'Antes', 'Después', 'Motivo', 'Detalle', 'Quién', 'Cuándo'))
        self.assertEqual(rows[1][0], product.name)
        self.assertEqual(rows[1][6], 'Merma/rotura')

    def test_export_requires_the_same_access_as_viewing(self):
        # Mismo gate que las 4 vistas ya prueban para JSON
        # (ReportsApiPermissionTests) — un cajero plano tampoco puede
        # exportar.
        self._auth(self.ctx['user'])  # handles_cash=True, sin can_authorize_exceptions
        response = self.client.get(
            '/api/v1/reports/sales-by-product/',
            {'date_from': self.today, 'date_to': self.today, 'export': 'xlsx'},
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_without_format_param_still_returns_json_as_before(self):
        # No debe romper el camino existente — format=xlsx es opt-in.
        self._auth(self.admin)
        response = self.client.get(
            '/api/v1/reports/sales-by-product/', {'date_from': self.today, 'date_to': self.today},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/json')


class ExportAuditLogTests(APITestCase):
    """Observación de sesión, punto 3: quién exportó, qué reporte, con qué
    filtros y cuándo — un registro en AuditLog, no una columna dentro del
    Excel."""

    def setUp(self):
        self.ctx = create_checkout_context()
        self.today = timezone.localdate()
        self.admin, _ = create_user_with_profile(
            'admin@donchuy.test', self.ctx['branch'], role=UserProfile.Role.ADMINISTRADOR,
        )

    def _auth(self, user):
        self.client.force_authenticate(user=user)

    def test_exporting_sales_by_product_logs_who_what_and_filters(self):
        self._auth(self.admin)
        self.client.get(
            '/api/v1/reports/sales-by-product/',
            {'date_from': self.today, 'date_to': self.today, 'export': 'xlsx', 'group_by': 'category'},
        )
        entry = AuditLog.objects.get(action='report.exported')
        self.assertEqual(entry.user, self.admin)
        self.assertEqual(entry.company, self.ctx['company'])
        self.assertEqual(entry.changes['report'], 'ventas-por-producto')
        self.assertEqual(entry.changes['group_by'], 'category')
        self.assertEqual(entry.changes['date_from'], str(self.today))
        self.assertIsNone(entry.changes['branch'])

    def test_exporting_with_a_branch_filter_logs_the_branch_name(self):
        self._auth(self.admin)
        self.client.get(
            '/api/v1/reports/inventory-valuation/',
            {'export': 'xlsx', 'branch': self.ctx['branch'].id},
        )
        entry = AuditLog.objects.get(action='report.exported')
        self.assertEqual(entry.changes['report'], 'valuacion-de-inventario')
        self.assertEqual(entry.changes['branch'], self.ctx['branch'].name)

    def test_exporting_cash_shift_detail_logs_which_shift(self):
        close_shift(shift=self.ctx['shift'], closing_user=self.ctx['user'], actual_closing_balance=Decimal('0'))
        self._auth(self.admin)
        self.client.get(
            '/api/v1/reports/cash-shift-detail/', {'shift': self.ctx['shift'].id, 'export': 'xlsx'},
        )
        entry = AuditLog.objects.get(action='report.exported')
        self.assertEqual(entry.changes['report'], 'cierre-de-turno-detallado')
        self.assertEqual(entry.changes['shift'], self.ctx['shift'].id)

    def test_viewing_json_without_exporting_does_not_log_anything(self):
        self._auth(self.admin)
        self.client.get(
            '/api/v1/reports/sales-by-product/', {'date_from': self.today, 'date_to': self.today},
        )
        self.assertFalse(AuditLog.objects.filter(action='report.exported').exists())
